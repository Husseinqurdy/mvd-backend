"""
Parses Excel (.xlsx) or CSV timetable files.
Expected columns: day, start_time, end_time, course_code, course_name,
                  venue_code, lecturer_name, program_code, year_of_study, group
"""
import re, csv, io
from dataclasses import dataclass
from datetime import time, datetime
from typing import Optional

try:
    import openpyxl
    HAS_XL = True
except ImportError:
    HAS_XL = False


@dataclass
class ExcelRow:
    day: int
    start_time: time
    end_time: time
    course_code: str
    course_name: str
    venue_code: str
    lecturer_name: str
    program_code: str
    year_of_study: int
    group: str
    is_cross: bool


DAY_MAP = {
    'monday': 1, 'mon': 1, 'tuesday': 2, 'tue': 2,
    'wednesday': 3, 'wed': 3, 'thursday': 4, 'thu': 4,
    'friday': 5, 'fri': 5, 'saturday': 6, 'sat': 6,
    '1': 1, '2': 2, '3': 3, '4': 4, '5': 5, '6': 6,
}

ALIASES = {
    'day':            ['day', 'day_of_week', 'weekday'],
    'start_time':     ['start_time', 'start', 'from'],
    'end_time':       ['end_time', 'end', 'to'],
    'course_code':    ['course_code', 'code', 'module_code'],
    'course_name':    ['course_name', 'name', 'subject', 'course'],
    'venue_code':     ['venue_code', 'venue', 'room'],
    'lecturer_name':  ['lecturer_name', 'lecturer', 'instructor'],
    'program_code':   ['program_code', 'program', 'programme'],
    'year_of_study':  ['year_of_study', 'year'],
    'group':          ['group', 'grp'],
    'is_cross':       ['is_cross', 'cross_cutting', 'cross'],
}

_alias_map = {a: f for f, al in ALIASES.items() for a in al}


def _to_time(val) -> Optional[time]:
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    m = re.match(r'^(\d{1,2}):(\d{2})', str(val).strip())
    return time(int(m.group(1)), int(m.group(2))) if m else None


def _build_row(d: dict) -> Optional[ExcelRow]:
    try:
        day = DAY_MAP.get(str(d.get('day', '')).strip().lower())
        if not day:
            return None
        t1 = _to_time(d.get('start_time', ''))
        t2 = _to_time(d.get('end_time', ''))
        if not t1 or not t2:
            return None
        code = str(d.get('course_code', '')).strip().upper()
        if not code:
            return None
        return ExcelRow(
            day=day, start_time=t1, end_time=t2,
            course_code=code,
            course_name=str(d.get('course_name', '')).strip(),
            venue_code=str(d.get('venue_code', '')).strip().upper(),
            lecturer_name=str(d.get('lecturer_name', '')).strip(),
            program_code=str(d.get('program_code', '')).strip().upper(),
            year_of_study=int(d.get('year_of_study', 1) or 1),
            group=str(d.get('group', '')).strip().upper(),
            is_cross=str(d.get('is_cross', 'false')).lower() in ('1', 'true', 'yes'),
        )
    except (ValueError, TypeError):
        return None


def parse_excel(source) -> list:
    if not HAS_XL:
        raise ImportError('openpyxl not installed. Run: pip install openpyxl')
    if isinstance(source, bytes):
        source = io.BytesIO(source)
    wb   = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in rows[0]]
    col_map = {i: _alias_map[h] for i, h in enumerate(headers) if h in _alias_map}
    result  = []
    for raw in rows[1:]:
        if all(v is None for v in raw):
            continue
        d = {field: raw[i] for i, field in col_map.items() if i < len(raw)}
        r = _build_row(d)
        if r:
            result.append(r)
    return result


def parse_csv(source) -> list:
    if isinstance(source, bytes):
        source = io.StringIO(source.decode('utf-8-sig'))
    elif hasattr(source, 'read'):
        content = source.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8-sig')
        source = io.StringIO(content)
    reader = csv.DictReader(source)
    result = []
    for raw in reader:
        d = {_alias_map.get(k.strip().lower().replace(' ', '_'), k): v for k, v in raw.items()}
        r = _build_row(d)
        if r:
            result.append(r)
    return result


def generate_template() -> bytes:
    if not HAS_XL:
        raise ImportError('openpyxl not installed')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Timetable'
    ws.append(['day', 'start_time', 'end_time', 'course_code', 'course_name',
               'venue_code', 'lecturer_name', 'program_code', 'year_of_study', 'group'])
    ws.append(['Monday', '07:30', '08:15', 'IT 6125', 'Computerized Accounting',
               'COMP. LAB 02', 'Nyambo N', 'DICT-Y1', '1', ''])
    ws2 = wb.create_sheet('Instructions')
    ws2.append(['Column', 'Example', 'Notes'])
    ws2.append(['day', 'Monday', 'Monday/Tuesday/Wednesday/Thursday/Friday/Saturday'])
    ws2.append(['start_time', '07:30', 'HH:MM format'])
    ws2.append(['end_time', '08:15', 'HH:MM format'])
    ws2.append(['course_code', 'IT 6125', 'As shown on timetable'])
    ws2.append(['course_name', 'Computerized Accounting', 'Full module name'])
    ws2.append(['venue_code', 'COMP. LAB 02', 'As shown on timetable'])
    ws2.append(['lecturer_name', 'Nyambo N', 'Surname + initial'])
    ws2.append(['program_code', 'DICT-Y1', 'Your program code'])
    ws2.append(['year_of_study', '1', '1, 2 or 3'])
    ws2.append(['group', 'GRP1', 'Leave blank if no group'])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
